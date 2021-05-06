import os
from pathlib import Path
import openpyxl


if not os.path.exists(Path.home() / 'TBFlash' / 'collection.xlsx'):
	if not os.path.exists(Path.home() / 'TBFlash'):
		os.mkdir(Path.home() / 'TBFlash')
	wb = openpyxl.Workbook()
	ws = wb.active
	ws['A1'] = 'ID'
	ws['B1'] = 'Question'
	ws['C1'] = 'Answer'
	wb.save(Path.home() / 'TBFlash' / 'collection.xlsx')

wb = openpyxl.load_workbook(Path.home() / 'TBFlash' / 'collection.xlsx')
ws = wb.active
collection = {}

class FlashCard:
	def __init__(self, cid, front, back):
		self.cid = cid
		self.front = front
		self.back = back

	def __str__(self):
		return str(self.front)

	def __repr__(self):
		return str(self.front)

	def print_card(self):
		print('Q:', self.front)
		print('A:', self.back)

def load_cards():
	if ws.max_row > 1:
		for i in range(2, ws.max_row+1):
			collection['card_'+str(i)] = FlashCard(ws.cell(row=i, column=1), ws.cell(row=i, column=2).value, ws.cell(row=i, column=3).value)

def print_menu():
	print()
	print("=== MENU ===")
	print("1. Add flashcard")
	print("2. Quiz")
	print("3. Show flashcards")
	print("9. Exit")
	print()

def add_card():
	print('=== NEW CARD ===')
	new_row = ws.max_row + 1
	question = input('Type question: ')
	answer = input('Type answer: ')
	ws.cell(row=new_row, column=1).value = get_new_cid()
	ws.cell(row=new_row, column=2).value = question
	ws.cell(row=new_row, column=3).value = answer
	wb.save(Path.home() / 'TBFlash' / 'collection.xlsx')
	print()
	load_cards()

def run_quiz():
	print('=== QUIZ - ', ws.max_row - 1, 'questions total ===')
	for i in range(2, ws.max_row+1):
		print('Question:', collection['card_'+str(i)].front)
		user_answer = input('Your answer: ')
		print('The correct answer is:', collection['card_'+str(i)].back)
		print()

def show_cards():
	print('=== SHOW CARDS ===')
	if ws.max_row <= 1:
		print('No cards to show.')
	print()
	for i in range(2, ws.max_row+1):
		collection['card_'+str(i)].print_card()
	print()

def get_new_cid():
	cid_list = []
	if ws.max_row > 1:
		for i in range(2, ws.max_row+1):
			cid_list.append(ws.cell(row=i, column=1).value)
		new_cid = max(cid_list) + 1
	elif ws.max_row == 1:
		new_cid = 1
	return new_cid

load_cards()

print_menu()
menu_choice = 0

while menu_choice != 9:
	menu_choice = int(input("Menu choice: "))
	print()
	if menu_choice == 1:
		add_card()
		print_menu()
	elif menu_choice == 2:
		run_quiz()
		print_menu()
	elif menu_choice == 3:
		show_cards()
		print_menu()
	elif menu_choice != 9:
		print("Invalid option, please try again.")
		print()